#utils
import openpyxl
from datetime import datetime
import os

#kivy
from kivy.app import App
from kivy.lang import Builder
from kivy.core.window import Window

#kivy utils
from kivy.uix.gridlayout import GridLayout
from kivy.uix.scrollview import ScrollView
from kivy.uix.label import Label
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.textinput import TextInput
from kivy.uix.button import Button
from kivy.uix.progressbar import ProgressBar
from kivy.utils import get_color_from_hex as getHex
from kivy.graphics import Color, Rectangle

KV = """
#:import utils kivy.utils
#:import os os
BoxLayout:
    orientation: 'horizontal'
    canvas.before:
        Color:
            rgba: utils.get_color_from_hex('#000000')
        Rectangle:
            size: self.size
            pos: self.pos
    
    spacing: 10
    padding: 10

    #LeftUpper
    BoxLayout:
        orientation:'vertical'
        spacing: 10
        padding: 10

        canvas.before:
            Color:
                rgba: utils.get_color_from_hex('#2e2e2e')
            Rectangle:
                size: self.size
                pos: self.pos

        GridLayout:
            cols: 1
            height: 
            canvas.before:
                Color:
                    rgba: utils.get_color_from_hex('#525252')
                Rectangle:
                    size: self.size
                    pos: self.pos

            Label:
                text: "日期"
                font_name: r'C:\Windows\Fonts\simhei.ttf'

            TextInput:
                id: date_input
                multiline: False
                hint_text: "输入日期"
                font_name: r'C:\Windows\Fonts\simhei.ttf'

            Label:
                text: "金额"
                font_name: r'C:\Windows\Fonts\simhei.ttf'

            TextInput:
                id: money_amount
                multiline: False
                hint_text: "输入金额"
                input_filter: 'float'
                font_name: r'C:\Windows\Fonts\simhei.ttf'
            
            Label:
                text: "用途"
                font_name: r'C:\Windows\Fonts\simhei.ttf'

            TextInput:
                id: money_usage
                multiline: True
                hint_text: "输入用途"
                font_name: r'C:\Windows\Fonts\simhei.ttf'
            
            Button:
                text: "提交"
                id: submit_button
                on_release: app.submitButtonClick()
                font_name: r'C:\Windows\Fonts\simhei.ttf'

        #LeftLower
        BoxLayout:
            orientation: 'vertical'
            canvas.before:
                Color:
                    rgba: utils.get_color_from_hex('#525252')
                Rectangle:
                    size: self.size
                    pos: self.pos
            Label: 
                text: "剩余钱数"
                size_hint_y: None
                height: 40
                font_name: r'C:\Windows\Fonts\simhei.ttf'
            
            BoxLayout:
                orientation: 'horizontal'
                size_hint_y: None
                height: 60
                BoxLayout:
                    orientation: 'horizontal'
                    Label:
                        text: "总限制金额:"
                        id: total_amount
                        size_hint_y: None
                        height: 40
                        font_name: r'C:\Windows\Fonts\simhei.ttf'

                    TextInput:
                        id: total_amount_input
                        multiline: False
                        hint_text: "输入限制金额"
                        size_hint_y: None
                        height: 40
                        input_filter: 'float'
                        on_text: app.onTextChange()
                        font_name: r'C:\Windows\Fonts\simhei.ttf'

                Label:
                    text: "总用量:"
                    id: total_usage
                    size_hint_y: None
                    height: 40
                    font_name: r'C:\Windows\Fonts\simhei.ttf'
            Label:
                text: "百分比"
                id: percentage
                size_hint_y: None
                height: 40
                font_name: r'C:\Windows\Fonts\simhei.ttf'

            ProgressBar:
                id: progressBar
                size_hint: (0.5, 1)
                max: 1
                pos_hint: {'center_x': 0.5, 'center_y': 0.5}

    #RightPart
    BoxLayout:
        orientation:'vertical'
        canvas.before:
            Color:
                rgba: utils.get_color_from_hex('#2e2e2e')
            Rectangle:
                size: self.size
                pos: self.pos
        spacing: 10
        padding: 10

        Label:
            text: "消费记录"
            #background_normal: ''
            #background_color: utils.get_color_from_hex('#00ccff')
            size_hint_y: None
            height: 100
            font_name: r'C:\Windows\Fonts\simhei.ttf'

        ScrollView:
            canvas.before:
                Color:
                    rgba: utils.get_color_from_hex('#525252')
                Rectangle:
                    size: self.size
                    pos: self.pos
            size_hint_x: 1
            BoxLayout:
                id: ScrollBoxLayoutContent
                orientation: 'vertical'
                size_hint_y: None
                height: self.minimum_height
                spacing: 10
                padding: 10
        Button:
            text: "保存变更"
            on_release: app.saveButtonClick()
            size_hint_y: None
            height: 60
            font_size: 40
            font_name: r'C:\Windows\Fonts\simhei.ttf'

"""

class Financial:
    def __init__(self):
        self.totalUsed = 0
        self.totalAmount = 0
        self.dailyExpenses = []
        self.currentSeq = 0
    
    class DailyExpense:
        def __init__(self, seq, amount, usage, date):
            self.seq = seq #int
            self.amount = amount #float
            self.usage = usage #string
            self.date = date #string

    def removeExpense(self, seqToDelete):
        expenseToRemove = None
        for expense in self.dailyExpenses:
            if expense.seq == seqToDelete:
                expenseToRemove = expense
                break
        maxSeq = 0
        if expenseToRemove:
            self.dailyExpenses.remove(expenseToRemove)  
            for i in range(0 , len(self.dailyExpenses)): 
                self.dailyExpenses[i].seq = i + 1
        self.currentSeq = len(self.dailyExpenses)
    #add a new daily expense
    def addDailyExpense(self, seq, amount, usage, date):
        dailyExpense = Financial.DailyExpense(seq, amount, usage, date)
        self.dailyExpenses.append(dailyExpense)
    
    def addDailyExpense(self, newExpense):
        self.dailyExpenses.append(newExpense)

    def calculatetotalUsed(self):
        self.totalUsed = 0
        for dailyExpense in self.dailyExpenses:
            self.totalUsed += dailyExpense.amount

class MyApp(App):
    def build(self):
        Window.size = (1200, 800)
        Window.resizable = False
        return Builder.load_string(KV)
    
    def on_start(self):
        #print("Starting app")
        self.initScrollContent()

    def initScrollContent(self):
        for logPart in financial.dailyExpenses:
            self.addNewExpenseWidget(logPart)
        financial.calculatetotalUsed()
        self.root.ids.total_amount_input.text = str(financial.totalAmount)
        self.root.ids.total_usage.text = f"Total Usage: {financial.totalUsed:.2f}"

    def updateScrollContent(self, newExpense): #newExpense: DailyExpense
        self.addNewExpenseWidget(newExpense)
        financial.calculatetotalUsed()
        self.root.ids.total_usage.text = f"Total Usage: {financial.totalUsed:.2f}"
        #result += f"{newExpense.seq}: {newExpense.date}, {newExpense.amount}, {newExpense.usage}\n"
        #self.root.ids.scrollViewContent.text = result
        pass

    def submitButtonClick(self):
        if not self.root.ids.money_amount.text.strip():
            return
        try:
            inputFieldFloat = float(self.root.ids.money_amount.text)
        except ValueError:
            return
        dateInput = self.root.ids.date_input.text
        if not dateInput.strip():
            dateInput = todayDate
        else:
            try:
                dateObj = datetime.strptime(dateInput, '%Y.%m.%d')
                dateInput = dateObj.strftime('%Y.%m.%d')
            except ValueError:
                return
        logNewPart = Financial.DailyExpense(financial.currentSeq+1, inputFieldFloat, self.root.ids.money_usage.text, dateInput)
        financial.currentSeq += 1
        financial.addDailyExpense(logNewPart)
        financial.calculatetotalUsed()
        self.updateScrollContent(logNewPart)
        self.updateProgressBar()
        pass

    def saveButtonClick(self):
        global sheet
        sheet.delete_rows(1, sheet.max_row)
        for expense in financial.dailyExpenses:
            sheet.append([expense.seq, expense.date, expense.amount, expense.usage])
        sheet['F1'] = financial.totalAmount
        log.save("sav.xlsx")

    def addNewExpenseWidget(self, newLog):
        newBox = BoxLayout(orientation='horizontal', size_hint_y=None, height=40) 
        with newBox.canvas.before:
            Color(*getHex("#808080"))
            newBox.rect = Rectangle(size=newBox.size, pos=newBox.pos)
        def updateRect(instance, value):
            instance.rect.size = instance.size
            instance.rect.pos = instance.pos
        newBox.bind(size=updateRect, pos=updateRect)
        newBox.add_widget(
            Label(text=f"{newLog.seq}: {newLog.date}, {newLog.amount}, {newLog.usage}", 
                size_hint_y=None, height=40, font_name=r'C:\Windows\Fonts\simhei.ttf')
        )
        newBox.add_widget(Button(text="删除", size_hint_x=None, width=70, 
            on_release=lambda x: self.deleteButtonClick(newLog.seq), font_name=r'C:\Windows\Fonts\simhei.ttf'))
        self.root.ids.ScrollBoxLayoutContent.add_widget(newBox)
        self.root.ids.ScrollBoxLayoutContent.height += newBox.height

    def deleteButtonClick(self, seqToDelete):
        financial.removeExpense(seqToDelete)
        self.root.ids.ScrollBoxLayoutContent.height = 0
        self.root.ids.ScrollBoxLayoutContent.clear_widgets()
        self.initScrollContent()
        self.updateProgressBar()
        pass

    def updateProgressBar(self):
        if financial.totalUsed != 0:
            try:
                inputFieldFloat = float(self.root.ids.total_amount_input.text)
                if inputFieldFloat <= 0:
                    return
            except ValueError:
                return
            financial.totalAmount = inputFieldFloat
            percent = financial.totalUsed / inputFieldFloat
            if percent > 1:
                percent = 1
            elif percent < 0:
                percent = 0
            self.root.ids.progressBar.value = percent
            remain = inputFieldFloat - financial.totalUsed
            self.root.ids.percentage.text = "已使用" + f"{int(percent*100)}%" + ", 剩余" + f"{remain:.2f}"
            #print(self.root.ids.progressBar.value)

    def onTextChange(self):
        self.updateProgressBar()

'''
Initialize everything, though it's not necessary to make part of
it a function but I JUST WANT IT BE ORGANIZED.
'''
def MainInitialize():
    global financial
    financial = Financial()
    global log
    if os.path.isfile("sav.xlsx"):
        log = openpyxl.load_workbook("sav.xlsx")
    else:
        log = openpyxl.Workbook()
        financial.totalAmount = 0
        log.save("sav.xlsx")
    global sheet
    sheet = log.active
    global todayDate
    todayDate = datetime.now().strftime("%Y.%m.%d")
    
    for row in sheet.iter_rows(values_only=True, min_row=1):
        if row is None or all(cell is None for cell in row):
            continue
        logPart = Financial.DailyExpense(row[0], row[2], row[3], row[1])
        financial.dailyExpenses.append(logPart)
        financial.currentSeq += 1
    try:
        inputFieldFloat = float(sheet['F1'].value)
        if inputFieldFloat <= 0:
            inputFieldFloat = 0
    except ValueError:
        inputFieldFloat = 0
    except TypeError:
        inputFieldFloat = 0
    financial.totalAmount = inputFieldFloat

if __name__ == '__main__':
    MainInitialize()
    MyApp().run()